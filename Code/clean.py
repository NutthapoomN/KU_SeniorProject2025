from openpyxl import load_workbook
import pandas as pd

wb = load_workbook("Excel/20_1_2026.xlsx")
ws = wb.active   # หรือ wb["ชื่อชีท"]
X = []
Y = []
for row in range(2, ws.max_row + 1):
    val = ws[f"C{row}"].value
    if val is not None:
        X.append(val)
for i in X:
    t= i
    seconds = t.hour*3600 + t.minute*60 + t.second
    Y.append(seconds)
print(len(X),len(Y))

start_row = 2  # ข้าม header

for i, value in enumerate(Y):
    ws.cell(row=start_row + i, column=1, value=value)

wb.save("Excel/20_1_2026.xlsx")
